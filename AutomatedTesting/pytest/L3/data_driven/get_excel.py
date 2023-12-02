# -*- coding: utf-8 -*-
# @Time : 2023/11/23 16:49
# @Author : 另外我
# @File : get_excel.py
# @Software: PyCharm


import openpyxl


def obt_book():
    # 获取工作簿
    book = openpyxl.load_workbook('123.xlsx')
    # 获取工作表sheet1
    sheet = book.active
    # 读取单个单元格--A1
    a_1 = sheet['A1'].value
    print(a_1)
    # 读取单个单元格（第三列，第三行）(column=列,row=行)==C3
    c_3 = sheet.cell(column=3, row=3).value
    print(c_3)
    # 读取多个连续单元格==a1到c3，输出为元组
    a1_c3 = sheet["A1":"C3"]
    print(type(a1_c3), a1_c3)
    new = list(a1_c3)
    print(type(new), new)


#
def test_excel():
    book = openpyxl.load_workbook('data/123.xlsx')
    sheet = book.active
    # cells = sheet["A1":"C3"]
    # print(cell)
    values = []
    for i in sheet:
        # 使用i变量遍历sheet的值
        cell = []
        for x in i:
            # 使用变量x遍历i的值（即sheet元组里面嵌套的三个元组）
            cell.append(x.value)
        values.append(cell)
        # return values
    print(values)


test_excel()
